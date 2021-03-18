#!/usr/bin/env python
# coding: utf-8
# Copyright 2021 Netbeam SAS
# Created by: Camilo Jimenez
""" Excel File class """
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from components.utils import CommonsUtils

# Init config Log
logging = CommonsUtils.setup_logging()

class ExcelFile:
    def __init__(self, src):
        self.src = src
        self.wb = Workbook()

    def set_sheet(self, sheet) :
        try:
            del self.wb['Sheet']
            self.wb.create_sheet(sheet)
        except Exception as error:
            logging.error('Error creando hoja en excel')  
            logging.error(f'Error: {error}')
            raise  

        self.ws = self.wb.get_sheet_by_name(sheet)

    def write_workbook(self,row_dest,column_dest,value):
        c = self.ws.cell(row = row_dest, column = column_dest)
        c.value = value

    def write_row(self, row):
        self.ws.append(row)

    def save_excel(self) :  
        self.wb.save(self.src)
    
    def create_table(self, table_name, range):
        self.tab = Table(displayName=table_name, ref=range)
        
        self.style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        self.tab.tableStyleInfo = self.style

        self.ws.add_table(self.tab, autofilter=0)
