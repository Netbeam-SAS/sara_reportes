#!/usr/bin/env python
# coding: utf-8
# Copyright 2021 Netbeam SAS
# Created by: Camilo Jimenez
""" Excel File class """
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from components.utils import CommonsUtils

# Init config Log
logging = CommonsUtils.setup_logging()

class ExcelFile:
    def __init__(self, src):
        self.src = src
        self.wb = Workbook()

    def set_sheet(self, sheet) :
        try:
            del self.wb['Sheet']
            self.wb.create_sheet(sheet)
        except Exception as error:
            logging.error('Error creando hoja en excel')  
            logging.error(f'Error: {error}')
            raise  

        self.ws = self.wb.get_sheet_by_name(sheet)

    def write_workbook(self,row_dest,column_dest,value):
        c = self.ws.cell(row = row_dest, column = column_dest)
        c.value = value

    def write_row(self, row):
        self.ws.append(row)
    
    def set_styles_file(self):
        column_widths = []        
        for row in self.ws.rows:            
            for i, cell in enumerate(row):
                cell_value = str(cell.value)
                if len(column_widths) > i:
                    if len(cell_value) > column_widths[i]:
                        column_widths[i] = len(cell_value)
                else:
                    column_widths += [len(cell_value)]

        for i, column_width in enumerate(column_widths):
            self.ws.column_dimensions[get_column_letter(i+1)].width = column_width + 2

    def save_excel(self, quantity):
        self.set_styles_file()
        self.create_table(table_name='tiempos', quantity=quantity)  
        self.wb.save(self.src)
    
    def create_table(self, table_name, quantity):
        range_table = f'A1:{get_column_letter(self.ws.max_column)}{quantity + 1}'
        table = Table(displayName=table_name, ref=range_table)
        
        table_style = TableStyleInfo(name="TableStyleLight11", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        table.tableStyleInfo = table_style

        self.ws.add_table(table)
    
